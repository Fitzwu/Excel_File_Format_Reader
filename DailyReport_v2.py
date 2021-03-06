from InsideFile import InsideFile, config, OwnSheet
import os
import json
import time
import datetime
import re
import sys,getopt
import csv
import xlwt
import xlrd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Fill

m_path = os.path.dirname(__file__)
m_jsonFile_path = os.path.join(m_path,"DailyReport.json")
try:
    jsonFile = open(m_jsonFile_path,'r' ,encoding='utf-8')
    pass
except:
    print("\n 请确认配置文件 DailyReport.json 存在")
    pass
m_jsonctx = json.load(jsonFile) #is a dict



'''
相关全局参数
'''
m_NewBook = Workbook()
m_NewSheet = m_NewBook.active
m_tmpdir = os.path.join(m_path,"tmp")
m_month = datetime.datetime.now().month
m_day = datetime.datetime.now().day
m_year = datetime.datetime.now().year
m_startday = datetime.datetime(m_year,1,1)
m_now = datetime.datetime(m_year, m_month, m_day)
m_src_LineNum = (m_now - m_startday).days + 4
# m_dst_LineNum = (m_now - m_startday).days + 4



# 得到当前文件下以日期为名的子文件夹
def getPath():
    TODAY = datetime.date.today() # year-month-day
    today = re.sub(r'\D',"","{}".format(TODAY)) # yearmonthday
    global m_path
    Path = os.path.join(m_path,today)
    return Path


# <std::map> 的查找功能
def stdMap(power_name):
    global m_jsonctx
    for i, power in enumerate(m_jsonctx["电站顺序"]):
        if power_name == power:
            return i
        pass
    pass
    return None

'''
Excel 读写
读 用的是 xlrd
写 用的是 openpyxl
'''

def ReadCell(sheet, LineNum, ColumnNum):
    return sheet.cell(LineNum - 1, ColumnNum - 1).value

def WriteCell(sheet, LineNum, ColumnNum, Value):
    return sheet.cell(row = LineNum, column = ColumnNum, value = Value)

def ReadLineValue(sheet, LineNum, startcol = 0, endcol = None):
    return sheet.row_values(LineNum-1, startcol, endcol)



def WriteRowValue(sheet, src_value, dst_line):
    for i, value in enumerate(src_value):
        sheet.cell(dst_line + 1, i + 1, value)
        
    # for rows in range(1,sheet.max_row+1):
    #     for cols in range(6,sheet.max_column+1):
    #         sheet.cell(rows,cols).fill = PatternFill("solid",fgColor="92D050")

    return


def ReadBlock(sheet, left_line, left_col, right_line, right_col):
    pass

def WriteBlock():
    pass


def SaveSingleFile(single_file_name, single_file_path, file_config, dst_sheet):
    book = xlrd.open_workbook(single_file_path)
    # m_book = xlrd.open_workbook(single_excel_path)
    
    sheet_conf = file_config.getOwnsheet()
    for conf in sheet_conf:
        try:
            sheet = book.sheet_by_name(conf.SheetName)
            pass
        except:
            print("\n",single_file_name,"中不存在名为 “", conf.SheetName ,"” 的工作表页!\n请确保各文件中存在对应名称的工作表页\n请注意表页名称不能有空格或其他无关字符！\n")
            os.system("pause")
            pass

        global m_src_LineNum
        # value = sheet.row_values(m_src_LineNum-1)
        tmplist = list()
        tmplist.clear()
        if not conf.range_list == None:
            for i in range(len(conf.range_list)):
                startcol = int(conf.range_list[i][0])
                endcol = int(conf.range_list[i][1]) + 1
                value = ReadLineValue(sheet, m_src_LineNum, startcol, endcol)
                # print(type(value))  
                # tmplist.insert(value)
                tmplist += value
                #TODO 拼接式读取
        else:
            value = ReadLineValue(sheet, m_src_LineNum)
            # print(type(value))
            tmplist += value
            

        WriteRowValue(dst_sheet, tmplist, dst_sheet.max_row+1)
        # dst_sheet.append(value)
    



List_File = []
def OpenAndReadFile(dirpath, startlevel):
    for insidefile in os.listdir(dirpath):
        insidefilePath = os.path.join(dirpath, insidefile)
        if os.path.isdir(insidefilePath):
            startlevel += 1 
            OpenAndReadFile(insidefilePath, startlevel)
        else:
            tmpfile = InsideFile(insidefile, startlevel, insidefilePath)
            List_File.append(tmpfile)
        

def Save2TmpXls():
    try:
        xlsx_list = os.listdir(getPath())
    except:
        print("\n当前目录不存在汇总文件夹！\n请确保当前路径下存在以当天日期为名的汇总文件夹 \n文件夹格式为 “20200601”\n")
        os.system("pause")
        pass
    
    '''
    os.path.isdir() //是目录
    os.path.isfile() //是文件
    '''
    
    LineNum = m_src_LineNum  
    m_tmpBook = Workbook()
    m_tmpSheet = m_tmpBook.active
    tmpXlsPath = os.path.join(m_tmpdir,"tmp.xlsx")
 
    global List_File
    List_File.clear()
    OpenAndReadFile(getPath(), 1)
    
    for itr in List_File:
        SaveSingleFile(itr.name, itr.File_Path(), itr.conf , m_tmpSheet)

    try:
        if not os._exists(tmpXlsPath):
            m_tmpBook.save(tmpXlsPath)
        else:
            os.remove(tmpXlsPath)
            m_tmpBook.save(tmpXlsPath)
    except:
        print("\n请先关闭当前文件：tmp.xlsx!\n")
        os.system("pause")




    
def Do():
    Save2TmpXls()

    # value_list = []
    # tmpXlsPath = os.path.join(m_tmpdir,"tmp.xlsx")
    # try:
    #     book = xlrd.open_workbook(tmpXlsPath)
    #     pass
    # except:
    #     print("\n找不到tmp.xlsx文件， 请查看tmp文件夹下是否存在相关文件!\n")
    #     os.system("pause")
    #     pass
    # sheet = book.sheet_by_index(0)
    # for row in range(sheet.nrows):
    #     value = sheet.row_values(row)
    #     value_list.append(sheet.row_values(row))
    #     i = stdMap(value[1])
    #     WriteRowValue(m_NewSheet, value, i)


        
Do()




    

