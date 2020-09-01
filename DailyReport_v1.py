#!/usr/bin/python3
import time
time_start = time.time()
import datetime
import re
import os
import sys,getopt
import csv
import xlwt
import xlrd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Fill

# requirements = ["openpyxl", "xlrd", "xlwt"]
# def check_requirement(package):
#     try:
#         exec("import {0}".format(package))
#     except ModuleNotFoundError:
#         print("\n缺少Python库:{0}\n正在安装……\n".format(package))
#         print("\n初次安装，时间可能较长，请耐心等待。\n")
#         os.system("pip3 install {0}".format(package))


        # inquiry = input("This script requires {0}. Do you want to install {0}? [y/n]".format(package))
        # while (inquiry != "y") and (inquiry != "n"):
        #     inquiry = input("This script requires {0}. Do you want to install {0}? [y/n]".format(package))
        # if inquiry == "y":
        #     print("Execute commands: pip install {0}".format(package))
            
        # else:
        #     print("{0} is missing, so the program exits!".format(package))
        #     exit(-1)
# for requirement in requirements:
#     check_requirement(requirement)



'''
相关全局参数
'''
m_NewBook = Workbook()
m_NewSheet = m_NewBook.active
m_tmpdir = os.path.join(os.path.dirname(__file__),"tmp")
m_month = datetime.datetime.now().month
m_day = datetime.datetime.now().day
m_year = datetime.datetime.now().year
m_startday = datetime.datetime(m_year,1,1)
m_now = datetime.datetime(m_year, m_month, m_day)
m_src_LineNum = (m_now - m_startday).days + 4
# m_dst_LineNum = (m_now - m_startday).days + 4

if not os.path.exists(m_tmpdir):
    m_tmpdir = os.makedirs(m_tmpdir)




'''
日期可变 但是文件夹不可变
'''
def setDate():
    date = input("请输入日期 (格式为2020-06-01) ：")
    if(date):
        input_year = date.split("-")[0]
        input_month = date.split("-")[1]
        input_day = date.split("-")[-1]
        day = int(input_day)
        month = int(input_month)
        year = int(input_year)
    tmp_now = datetime.datetime(year, month, day)
    global m_src_LineNum
    m_src_LineNum = (tmp_now - m_startday).days + 4
    

'''
得到当前文件下日期为名的子文件夹
'''
def getPath():
    TODAY = datetime.date.today() # year-month-day
    today = re.sub(r'\D',"","{}".format(TODAY)) # yearmonthday
    tmpPath = os.path.dirname(__file__)
    Path = os.path.join(tmpPath,today)
    return Path

'''
Excel 读写
'''

def ReadCell(sheet, LineNum, ColumnNum):
    return sheet.cell(LineNum - 1, ColumnNum - 1).value

def WriteCell(sheet, LineNum, ColumnNum, Value):
    return sheet.cell(row = LineNum, column = ColumnNum, value = Value)

def ReadLineValue(sheet, LineNum):
    return sheet.row_values(LineNum-1)

def WriteRowValue(sheet, src_value, dst_line):
    for i, value in enumerate(src_value):
        sheet.cell(dst_line + 1, i + 1, value)
        
    for rows in range(1,sheet.max_row+1):
        for cols in range(6,sheet.max_column+1):
            sheet.cell(rows,cols).fill = PatternFill("solid",fgColor="92D050")

    return

def Save2TmpXls():
    try:
        xlsx_list = os.listdir(getPath())
    except:
        print("\n当前目录不存在汇总文件夹！\n请确保当前路径下存在以当天日期为名的汇总文件夹 \n文件夹格式为 “20200601”\n")
        os.system("pause")
        pass
    
    '''
    os.path.isdir() //是目录
    os.path.isfile() //是文件
    '''
    
    LineNum = m_src_LineNum  
    m_tmpBook = Workbook()
    m_tmpSheet = m_tmpBook.active
    tmpXlsPath = os.path.join(m_tmpdir,"tmp.xlsx")
 
    for i, single_excel_file in enumerate(xlsx_list):
        single_excel_path = os.path.join(getPath(),single_excel_file)
        if not os.path.isfile(single_excel_path):
            tmpDir_QingHaiBeiXiang = single_excel_path 
            QHBX_excel = os.listdir(tmpDir_QingHaiBeiXiang)
            for i, QHBX_excel_file in enumerate(QHBX_excel):
                if (i > 1):
                    print("青海文件夹中文件数量不正确！\n")
                    exit(-1)
                else:
                    QHBX_excel_Path = os.path.join(tmpDir_QingHaiBeiXiang, QHBX_excel_file)
                    SaveQHBXFile(QHBX_excel_file, QHBX_excel_Path, m_tmpSheet)

        else:
            SaveSingleFile(single_excel_file, single_excel_path, m_tmpSheet)

    try:
        if not os._exists(tmpXlsPath):
            m_tmpBook.save(tmpXlsPath)
        else:
            os.remove(tmpXlsPath)
            m_tmpBook.save(tmpXlsPath)
    except:
        print("\n请先关闭当前文件：tmp.xlsx!\n")
        os.system("pause")




def SaveSingleFile(single_file_name, single_file_path ,dst_sheet):
    book = xlrd.open_workbook(single_file_path)
    # m_book = xlrd.open_workbook(single_excel_path)
    try:
        sheet = book.sheet_by_name('每日信息汇总页')
        pass
    except:
        print("\n",single_file_name,"中不存在名为 “每日信息汇总页” 的工作表页!\n请确保各文件中存在对应名称的工作表页\n请注意表页名称不能有空格或其他无关字符！\n")
        os.system("pause")
        pass
    global m_src_LineNum
    value = sheet.row_values(m_src_LineNum-1)
    dst_sheet.append(value)


def SaveQHBXFile(single_file_name, single_file_path ,dst_sheet):
    book = xlrd.open_workbook(single_file_path)
    # m_book = xlrd.open_workbook(single_excel_path)
    try:
        sheet1 = book.sheet_by_name('一电厂')
        sheet2 = book.sheet_by_name('二电厂')
        sheet3 = book.sheet_by_name('三电厂')
        pass
    except:
        print("\n",single_file_name,"中不存在名为 “一电厂/二电厂/三电厂” 的工作表页!\n请确保各文件中存在对应名称的工作表页\n请注意表页名称不能有空格或其他无关字符！\n")
        os.system("pause")
        pass
    global m_src_LineNum
    value1 = sheet1.row_values(m_src_LineNum-1)
    value2 = sheet2.row_values(m_src_LineNum-1)
    value3 = sheet3.row_values(m_src_LineNum-1)
    dst_sheet.append(value1)
    dst_sheet.append(value2)
    dst_sheet.append(value3)


'''
保存文件
'''
def saveFile():
    FileName = "{}年日报({}月{}日).xlsx".format(m_year, m_month, m_day)
    FilePath = os.path.join(os.path.dirname(__file__),FileName)
    try:
        if(os.path.exists(FilePath)):
            os.remove(FilePath)
            m_NewBook.save(FilePath)
        else:
            m_NewBook.save(FilePath)
    except:
        print("\n请先关闭当前文件:",FileName,"!\n")
        os.system("pause")

    try:
        if(os.path.exists(os.path.join(m_tmpdir,"tmp.xlsx"))):
            os.remove(os.path.join(m_tmpdir,"tmp.xlsx"))
        pass
    except:
        print("\n请先关闭当前文件：tmp.xlsx!\n")
        os.system("pause")
        pass

# <std::map> 的功能
def stdMap(power_name):
    Path = os.path.dirname(__file__)
    pp = os.path.join(Path,"电站顺序.txt")
    try:
        f =  open(pp,'r',encoding='utf8')
        pass
    except IOError as identifier:
        print("\n找不到文件：“电站顺序.txt” \n请确保当前文件夹下存在该文件！\n")
        os.system("pause")
        pass
    
    order_list = csv.reader(f)
    for i, power in enumerate(order_list):
        if power_name == power[0]:
            return i
        pass
    pass
    return None


def exceptation():
    pass




def Do():
    Save2TmpXls()
    value_list = []
    tmpXlsPath = os.path.join(m_tmpdir,"tmp.xlsx")
    try:
        book = xlrd.open_workbook(tmpXlsPath)
        pass
    except:
        print("\n找不到tmp.xlsx文件， 请查看tmp文件夹下是否存在相关文件!\n")
        os.system("pause")
        pass
    sheet = book.sheet_by_index(0)
    for row in range(sheet.nrows):
        value = sheet.row_values(row)
        value_list.append(sheet.row_values(row))
        i = stdMap(value[1])
        WriteRowValue(m_NewSheet, value, i)

def Finalcheck():
    FileName = "{}年日报({}月{}日).xlsx".format(m_year, m_month, m_day)
    FilePath = os.path.join(os.path.dirname(__file__),FileName)
    if not os._exists(FilePath): 
        Do()
        saveFile()
    else:
        return





def CheckInput(argv):
   try:
      opts, args = getopt.getopt(argv,"d")
   except getopt.GetoptError:
      pass

   for opt, arg in opts:
      if opt == '-d':
        setDate()

if __name__ == "__main__":
   CheckInput(sys.argv[1:])


Do()
saveFile()
Finalcheck()
time_end = time.time()
print("\n本次运行时间：%2f秒\n"%(time_end-time_start))
os.system("pause")
